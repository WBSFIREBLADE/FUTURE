import openpyxl
from io import BytesIO
from openai import OpenAI
from django.conf import settings
import json
import re
from datetime import datetime

client = OpenAI(api_key=settings.OPENAI_API_KEY)


def validate_excel_upload(file):
    errors = []

    # --- Layer 1: Load workbook ---
    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
    except Exception:
        return {
            "valid": False,
            "errors": [{"message": "Could not read the file. Make sure it is a valid .xlsx Excel file."}]
        }

    # --- Layer 2: Check header row ---
    header_row = [cell.value for cell in ws[1]]

    if not any(header_row):
        return {
            "valid": False,
            "errors": [{"message": "The file has no column headers in the first row."}]
        }

    for col_idx, header in enumerate(header_row, start=1):
        if header is None or str(header).strip() == "":
            errors.append({
                "row": 1,
                "column_index": col_idx,
                "message": f"Column {col_idx} has no header name"
            })

    col_names = [
        str(h).strip() if h and str(h).strip() != "" else f"Column {i+1}"
        for i, h in enumerate(header_row)
    ]

    # --- Layer 3: Check no data cell is empty ---
    all_rows = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        for col_idx, cell_value in enumerate(row):
            if cell_value is None or str(cell_value).strip() == "":
                errors.append({
                    "row": row_num,
                    "column": col_names[col_idx],
                    "column_index": col_idx + 1,
                    "message": f"Row {row_num} → '{col_names[col_idx]}' is empty"
                })

        all_rows.append({"row_num": row_num, "values": list(row)})

    # --- Layer 4: AI detects schema from sample only ---
    # only run if no empty cell errors
    if not errors:

        # send only first 10 rows to AI — enough to detect pattern
        sample_rows = all_rows[:10]

        # AI detects schema from sample
        ai_result = detect_schema_with_ai(col_names, sample_rows)

        if ai_result:
            schema = ai_result.get("column_types", {})
            ai_errors = ai_result.get("type_errors", [])

            # 🔥 ADD THIS LINE
            errors.extend(ai_errors)

            # run local validation also (optional but recommended)
            type_errors = validate_all_rows_locally(col_names, all_rows, schema)
            errors.extend(type_errors)

    if errors:
        return {"valid": False, "errors": errors}

    return {"valid": True, "schema": schema, "column_names": col_names}  # return schema and column names for table creation


def detect_schema_with_ai(col_names, sample_rows):
    # prepare sample data
    rows_preview = []
    for r in sample_rows:
        row_dict = {}
        for idx, col in enumerate(col_names):
            row_dict[col] = str(r["values"][idx]) if r["values"][idx] is not None else ""
        rows_preview.append(row_dict)

    prompt = f"""
You are a data validation assistant.

I have an Excel file with the following columns:
{json.dumps(col_names)}

Here are sample rows from the file:
{json.dumps(rows_preview, indent=2)}

Your tasks:

1. For each column, analyze all sample values.
2. Determine the data type of the column based on the MAJORITY of values.
3. Allowed data types:
   - INTEGER
   - DECIMAL
   - DATE
   - EMAIL
   - TEXT

4. If most values in a column are integers, the column type should be INTEGER.
   Even if some values do not match, still choose the majority type.

5. Identify rows where values do NOT match the detected majority data type.

Response rules:
- Respond ONLY in valid JSON.
- Do NOT include any explanation or extra text.

Return output in this exact format:

{{
  "column_types": {{
    "column_name": "detected_type"
  }},
  "type_errors": [
    {{
      "row": 2,
      "column": "column_name",
      "expected": "INTEGER",
      "found": "abc",
      "message": "Row 2 → 'column_name' should be INTEGER but found 'abc'"
    }}
  ]
}}

If there are no type errors, return an empty array for "type_errors".
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
        )

        content = response.choices[0].message.content.strip()
        return json.loads(content)

    except Exception:
        return None  # if AI fails skip type checking


def validate_all_rows_locally(col_names, all_rows, schema):
    errors = []

    email_pattern = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')
    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]

    for row_data in all_rows:
        row_num = row_data["row_num"]
        values = row_data["values"]

        for col_idx, col_name in enumerate(col_names):
            value = values[col_idx]
            expected_type = schema.get(col_name)

            if value is None or str(value).strip() == "":
                continue  # already caught in Layer 3

            val = str(value).strip()

            # check based on expected type
            if expected_type == "INTEGER":
                try:
                    int(val)
                except ValueError:
                    errors.append({
                        "row": row_num,
                        "column": col_name,
                        "column_index": col_idx + 1,
                        "expected": "INTEGER",
                        "found": val,
                        "message": f"Row {row_num} → '{col_name}' should be a whole number but found '{val}'"
                    })

            elif expected_type == "DECIMAL":
                try:
                    float(val)
                except ValueError:
                    errors.append({
                        "row": row_num,
                        "column": col_name,
                        "column_index": col_idx + 1,
                        "expected": "DECIMAL",
                        "found": val,
                        "message": f"Row {row_num} → '{col_name}' should be a number but found '{val}'"
                    })

            elif expected_type == "EMAIL":
                if not email_pattern.match(val):
                    errors.append({
                        "row": row_num,
                        "column": col_name,
                        "column_index": col_idx + 1,
                        "expected": "EMAIL",
                        "found": val,
                        "message": f"Row {row_num} → '{col_name}' should be a valid email but found '{val}'"
                    })

            elif expected_type == "DATE":
                parsed = False
                for fmt in date_formats:
                    try:
                        datetime.strptime(val, fmt)
                        parsed = True
                        break
                    except ValueError:
                        continue
                if not parsed:
                    errors.append({
                        "row": row_num,
                        "column": col_name,
                        "column_index": col_idx + 1,
                        "expected": "DATE",
                        "found": val,
                        "message": f"Row {row_num} → '{col_name}' should be a date but found '{val}'"
                    })

    return errors